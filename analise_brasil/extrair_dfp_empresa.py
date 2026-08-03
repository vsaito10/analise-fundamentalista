"""
Extrai dados históricos (2017-2024) de demonstrativos financeiros DFP de uma
empresa específica a partir dos arquivos consolidados da B3 e salva em Excel.

Demonstrativos suportados: DRE, BPA, BPP (e qualquer outro disponível).

Uso:
    python extrair_dfp_empresa.py

Ajuste as variáveis EMPRESA e OUTPUT_DIR abaixo para mudar a empresa ou destino.
"""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# ── Configuração ──────────────────────────────────────────────────────────────
PARQUET_DIR = Path("C:/B3/historico-arquivos/dfp-b3/dfp-cia-aberta/consolidados")
EMPRESA     = "WEG"
OUTPUT_DIR  = Path("C:/B3/analise-fundamentalista/analise_brasil")

# Demonstrativos a extrair:
#   (sigla, nome_aba, título, filtro_contas | None, keywords_extra | None)
#
# filtro_contas  — lista de CD_CONTA exatos a incluir (None = usar filtro de profundidade)
# keywords_extra — palavras-chave buscadas em DS_CONTA para incluir linhas extras além
#                  do filtro de profundidade (útil quando o código varia por empresa)
DEMONSTRATIVOS = [
    (
        "DRE",
        "DRE",
        "DRE Consolidada",
        [
            "3.01", "3.02", "3.03", "3.04", "3.05",
            "3.06", "3.06.01", "3.06.02",
            "3.07", "3.08", "3.08.01", "3.08.02",
            "3.09", "3.10", "3.11", "3.11.01",
        ],
        None,
    ),
    ("BPA",    "BPA",    "Balanço Patrimonial Ativo",              None, None),
    ("BPP",    "BPP",    "Balanço Patrimonial Passivo",            None, None),
    ("DFC_MI", "DFC_MI", "Demonstração do Fluxo de Caixa (Ind.)", None, ["deprecia"]),
]
# ─────────────────────────────────────────────────────────────────────────────


def _formatar_sheet(ws, titulo: str, empresa: str, escala: str) -> None:
    """Aplica formatação à worksheet já preenchida."""
    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 14   # Código
    ws.column_dimensions["B"].width = 55   # Descrição
    for i, col in enumerate(ws.iter_cols(min_col=3, max_col=ws.max_column)):
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = 16

    # Inserir 2 linhas de cabeçalho informativo no topo
    ws.insert_rows(1, 2)
    ws["A1"] = f"{titulo} — {empresa}"
    ws["A2"] = f"Valores em {escala} (BRL)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(italic=True, size=10)

    # Estilizar linha de cabeçalho das colunas (anos) — linha 3
    header_row = 3
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Zebra + formato numérico nas linhas de dados
    fill_even = PatternFill("solid", fgColor="DCE6F1")
    data_start = header_row + 1
    for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row)):
        fill = fill_even if i % 2 == 0 else None
        for j, cell in enumerate(row):
            if fill:
                cell.fill = fill
            if j >= 2:  # colunas numéricas (anos)
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    # Congelar painel após cabeçalho + 2 primeiras colunas
    ws.freeze_panes = ws.cell(row=data_start, column=3)


def _coalescer_linhas_keyword(pivot: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    """
    Para cada keyword, se houver mais de uma linha correspondente no pivot,
    une-as em uma única linha por coalesce (primeiro valor não-nulo por coluna).
    Usa o índice (Código, Descrição) da linha com mais valores preenchidos.
    Linhas que se complementam (períodos distintos, sem sobreposição) são o caso típico.
    """
    for kw in keywords:
        descricoes = pivot.index.get_level_values("Descrição").str.lower()
        mask = descricoes.str.contains(kw.lower(), na=False)
        matches = pivot[mask]

        if len(matches) <= 1:
            continue

        # Coalesce: primeiro valor não-nulo por coluna
        merged = matches.apply(
            lambda col: col.dropna().iloc[0] if col.notna().any() else float("nan")
        )

        # Manter o índice (Código, Descrição) da linha com mais dados preenchidos
        best_idx = matches.notna().sum(axis=1).idxmax()

        merged_row = pd.DataFrame(
            [merged.values],
            index=pd.MultiIndex.from_tuples([best_idx], names=pivot.index.names),
            columns=pivot.columns,
        )

        pivot = pd.concat([pivot[~mask], merged_row]).sort_index(level="Código")
        print(f"    Coalesce '{kw}': {len(matches)} linhas → 1  {[idx[1] for idx in matches.index]}")

    return pivot


def extrair_demonstrativo(
    sigla: str,
    empresa: str,
    parquet_dir: Path,
    contas_filtro: list[str] | None = None,
    keywords_extra: list[str] | None = None,
) -> tuple[pd.DataFrame, str, str]:
    """Lê o parquet de um demonstrativo e retorna (pivot, nome_empresa, escala)."""
    path = parquet_dir / f"dfp_cia_aberta_{sigla}_con_2017_2024.parquet"
    print(f"  Lendo {path.name} ...")
    df = pd.read_parquet(path)

    mask = df["DENOM_CIA"].str.contains(empresa, case=False, na=False)
    df_emp = df[mask].copy()
    if df_emp.empty:
        raise ValueError(f"Nenhuma empresa encontrada com o nome '{empresa}' em {path.name}")

    nome_empresa = df_emp["DENOM_CIA"].iloc[0]
    escala = df_emp["ESCALA_MOEDA"].iloc[0]

    # Apenas exercício corrente (ÚLTIMO) + versão mais recente por período
    df_emp = df_emp[df_emp["ORDEM_EXERC"].str.strip() == "ÚLTIMO"]
    df_emp = (
        df_emp.sort_values("VERSAO")
        .groupby(["DT_REFER", "CD_CONTA"], as_index=False)
        .last()
    )

    df_emp["ANO"] = pd.to_datetime(df_emp["DT_REFER"]).dt.year

    pivot = df_emp.pivot_table(
        index=["CD_CONTA", "DS_CONTA"],
        columns="ANO",
        values="VL_CONTA",
        aggfunc="first",
    )
    pivot.index.names = ["Código", "Descrição"]
    pivot.columns.name = None
    pivot = pivot.sort_index(level="Código")

    if contas_filtro:
        pivot = pivot[pivot.index.get_level_values("Código").isin(contas_filtro)]
    else:
        # Remover linhas onde todos os valores são nulos ou zero
        mask_valido = ~((pivot.isna() | (pivot == 0)).all(axis=1))
        # Manter apenas níveis 1, 2 e 3 (ex.: "1", "1.01", "1.01.01")
        mask_depth = pivot.index.get_level_values("Código").str.count(r"\.") <= 2
        # Incluir linhas extras por palavra-chave na descrição (independente do nível)
        if keywords_extra:
            descricoes = pivot.index.get_level_values("Descrição").str.lower()
            mask_kw = pd.Series(False, index=pivot.index)
            for kw in keywords_extra:
                mask_kw |= descricoes.str.contains(kw.lower(), na=False)
            pivot = pivot[mask_valido & (mask_depth | mask_kw)]
        else:
            pivot = pivot[mask_valido & mask_depth]

    if keywords_extra:
        pivot = _coalescer_linhas_keyword(pivot, keywords_extra)

    return pivot, nome_empresa, escala


def gerar_excel(empresa: str, output_dir: Path) -> None:
    empresa_slug = None

    for sigla, nome_aba, titulo, contas_filtro, keywords_extra in DEMONSTRATIVOS:
        print(f"\n[{sigla}]")
        pivot, nome_empresa, escala = extrair_demonstrativo(
            sigla, empresa, PARQUET_DIR, contas_filtro, keywords_extra
        )

        if empresa_slug is None:
            empresa_slug = nome_empresa.replace(" ", "_").replace("/", "-")

        output_path = output_dir / f"{sigla}_{empresa_slug}_2017_2024.xlsx"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pivot.to_excel(writer, sheet_name=nome_aba)
            _formatar_sheet(writer.sheets[nome_aba], titulo, nome_empresa, escala)

        print(f"  Salvo em: {output_path}")
        print(f"  Linhas (contas): {len(pivot)}  |  Anos: {list(pivot.columns)}")


if __name__ == "__main__":
    gerar_excel(EMPRESA, OUTPUT_DIR)
