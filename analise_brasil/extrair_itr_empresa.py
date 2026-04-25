"""
Extrai dados historicos trimestrais (2020-2025) de demonstrativos ITR de uma
empresa especifica a partir dos arquivos consolidados da B3 e salva em Excel.

Demonstrativos suportados: DRE, BPA, BPP (e qualquer outro disponivel).

Uso:
    python extrair_itr_empresa.py

Ajuste as variaveis EMPRESA e OUTPUT_DIR abaixo para mudar a empresa ou destino.
"""

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# Configuracao
PARQUET_DIR = Path("C:/B3/historico-arquivos/itr-b3/itr-cia-aberta/consolidados")
EMPRESA = "WEG"
OUTPUT_DIR = Path("C:/B3/analise-fundamentalista/analise_brasil")
PERIODO_ARQUIVOS = "2020_2025"

# Demonstrativos a extrair:
#   (sigla, nome_aba, titulo, filtro_contas | None, keywords_extra | None)
#
# filtro_contas  - lista de CD_CONTA exatos a incluir (None = usar filtro de profundidade)
# keywords_extra - palavras-chave buscadas em DS_CONTA para incluir linhas extras alem
#                  do filtro de profundidade (util quando o codigo varia por empresa)
DEMONSTRATIVOS = [
    (
        "DRE",
        "DRE",
        "DRE Consolidada Trimestral",
        [
            "3.01", "3.02", "3.03", "3.04", "3.05",
            "3.06", "3.06.01", "3.06.02",
            "3.07", "3.08", "3.08.01", "3.08.02",
            "3.09", "3.10", "3.11", "3.11.01",
        ],
        None,
    ),
    ("BPA", "BPA", "Balanco Patrimonial Ativo", None, None),
    ("BPP", "BPP", "Balanco Patrimonial Passivo", None, None),
    ("DFC_MI", "DFC_MI", "Demonstracao do Fluxo de Caixa (Ind.)", None, ["deprecia"]),
]


def _formatar_sheet(ws, titulo: str, empresa: str, escala: str) -> None:
    """Aplica formatacao a worksheet ja preenchida."""
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    for col in ws.iter_cols(min_col=3, max_col=ws.max_column):
        letra = col[0].column_letter
        ws.column_dimensions[letra].width = 16

    ws.insert_rows(1, 2)
    ws["A1"] = f"{titulo} - {empresa}"
    ws["A2"] = f"Valores em {escala} (BRL)"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"].font = Font(italic=True, size=10)

    header_row = 3
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    fill_even = PatternFill("solid", fgColor="DCE6F1")
    data_start = header_row + 1
    for i, row in enumerate(ws.iter_rows(min_row=data_start, max_row=ws.max_row)):
        fill = fill_even if i % 2 == 0 else None
        for j, cell in enumerate(row):
            if fill:
                cell.fill = fill
            if j >= 2:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=data_start, column=3)


def _coalescer_linhas_keyword(pivot: pd.DataFrame, keywords: list[str]) -> pd.DataFrame:
    """
    Para cada keyword, se houver mais de uma linha correspondente no pivot,
    une-as em uma unica linha por coalesce (primeiro valor nao-nulo por coluna).
    """
    for kw in keywords:
        descricoes = pivot.index.get_level_values("Descricao").str.lower()
        mask = descricoes.str.contains(kw.lower(), na=False)
        matches = pivot[mask]

        if len(matches) <= 1:
            continue

        merged = matches.apply(
            lambda col: col.dropna().iloc[0] if col.notna().any() else float("nan")
        )
        best_idx = matches.notna().sum(axis=1).idxmax()

        merged_row = pd.DataFrame(
            [merged.values],
            index=pd.MultiIndex.from_tuples([best_idx], names=pivot.index.names),
            columns=pivot.columns,
        )

        pivot = pd.concat([pivot[~mask], merged_row]).sort_index(level="Codigo")
        print(f"    Coalesce '{kw}': {len(matches)} linhas -> 1")

    return pivot


def _rotulo_trimestre(dt_refer: pd.Series) -> pd.Series:
    datas = pd.to_datetime(dt_refer)
    trimestre = datas.dt.quarter.astype(str)
    ano = datas.dt.year.astype(str)
    return ano + "Q" + trimestre


def extrair_demonstrativo(
    sigla: str,
    empresa: str,
    parquet_dir: Path,
    contas_filtro: list[str] | None = None,
    keywords_extra: list[str] | None = None,
) -> tuple[pd.DataFrame, str, str]:
    """Le o parquet de um demonstrativo ITR e retorna (pivot, nome_empresa, escala)."""
    path = parquet_dir / f"itr_cia_aberta_{sigla}_con_{PERIODO_ARQUIVOS}.parquet"
    print(f"  Lendo {path.name} ...")
    df = pd.read_parquet(path)

    mask = df["DENOM_CIA"].str.contains(empresa, case=False, na=False)
    df_emp = df[mask].copy()
    if df_emp.empty:
        raise ValueError(f"Nenhuma empresa encontrada com o nome '{empresa}' em {path.name}")

    nome_empresa = df_emp["DENOM_CIA"].iloc[0]
    escala = df_emp["ESCALA_MOEDA"].iloc[0]

    # Apenas periodo corrente (ULTIMO) + versao mais recente por trimestre/conta.
    df_emp = df_emp[df_emp["ORDEM_EXERC"].str.strip() == "ÚLTIMO"]
    df_emp = (
        df_emp.sort_values("VERSAO")
        .groupby(["DT_REFER", "CD_CONTA"], as_index=False)
        .last()
    )

    df_emp["TRIMESTRE"] = _rotulo_trimestre(df_emp["DT_REFER"])

    pivot = df_emp.pivot_table(
        index=["CD_CONTA", "DS_CONTA"],
        columns="TRIMESTRE",
        values="VL_CONTA",
        aggfunc="first",
    )
    pivot.index.names = ["Codigo", "Descricao"]
    pivot.columns.name = None
    pivot = pivot.reindex(sorted(pivot.columns), axis=1)
    pivot = pivot.sort_index(level="Codigo")

    if contas_filtro:
        pivot = pivot[pivot.index.get_level_values("Codigo").isin(contas_filtro)]
    else:
        mask_valido = ~((pivot.isna() | (pivot == 0)).all(axis=1))
        mask_depth = pivot.index.get_level_values("Codigo").str.count(r"\.") <= 2
        if keywords_extra:
            descricoes = pivot.index.get_level_values("Descricao").str.lower()
            mask_kw = pd.Series(False, index=pivot.index)
            for kw in keywords_extra:
                mask_kw |= descricoes.str.contains(kw.lower(), na=False)
            pivot = pivot[mask_valido & (mask_depth | mask_kw)]
        else:
            pivot = pivot[mask_valido & mask_depth]

    if keywords_extra:
        pivot = _coalescer_linhas_keyword(pivot, keywords_extra)

    return pivot, nome_empresa, escala


def gerar_excel(empresa: str, output_dir: Path) -> None:
    empresa_slug = None

    for sigla, nome_aba, titulo, contas_filtro, keywords_extra in DEMONSTRATIVOS:
        print(f"\n[{sigla}]")
        pivot, nome_empresa, escala = extrair_demonstrativo(
            sigla, empresa, PARQUET_DIR, contas_filtro, keywords_extra
        )

        if empresa_slug is None:
            empresa_slug = nome_empresa.replace(" ", "_").replace("/", "-")

        output_path = output_dir / f"{sigla}_{empresa_slug}_ITR_{PERIODO_ARQUIVOS}.xlsx"

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            pivot.to_excel(writer, sheet_name=nome_aba)
            _formatar_sheet(writer.sheets[nome_aba], titulo, nome_empresa, escala)

        print(f"  Salvo em: {output_path}")
        print(f"  Linhas (contas): {len(pivot)}  |  Trimestres: {list(pivot.columns)}")


if __name__ == "__main__":
    gerar_excel(EMPRESA, OUTPUT_DIR)
